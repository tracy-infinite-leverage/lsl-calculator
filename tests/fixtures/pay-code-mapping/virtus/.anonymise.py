#!/usr/bin/env python3
"""
Anonymise Virtus Health fixture sources before commit to the repo.

Source path: ~/Downloads/Virtus Health - LSL calculation/Sample run/
Output path: tests/fixtures/pay-code-mapping/virtus/

What we strip:
  - `Employee DOB` column in PositionHistory CSV and Sheet2 of the workbook.

What we preserve:
  - Employee IDs (numeric or short alphanumeric; not directly identifying).
  - All other columns: pay codes, dates, classifications, state names,
    employment-type prefixes, LSL Instrument labels, cohort strings. These are
    the canonical surfaces the E5.3 mapping + normalisation layer exercises.

Re-run after any source-file refresh. Outputs are byte-identical modulo
Excel encoding quirks (openpyxl rewrites formulas to literal values when the
workbook is saved without `data_only=False` — we preserve formulas explicitly).

Usage:
    python3 tests/fixtures/pay-code-mapping/virtus/.anonymise.py
"""

import csv
import os
import sys
from pathlib import Path

import openpyxl

SRC = Path(os.path.expanduser("~/Downloads/Virtus Health - LSL calculation/Sample run"))
OUT = Path(__file__).parent

if not SRC.exists():
    sys.stderr.write(
        f"Source folder not found: {SRC}\n"
        "Tracy's local copy of the Virtus sample files must be present at this path.\n"
    )
    sys.exit(1)


def anonymise_csv(src_name: str, dst_name: str, dob_columns: list[str]) -> None:
    """Copy CSV from src → dst, blanking any column whose header matches dob_columns."""
    src_path = SRC / src_name
    dst_path = OUT / dst_name
    with open(src_path, newline="", encoding="utf-8") as fin:
        reader = csv.reader(fin)
        rows = list(reader)
    if not rows:
        sys.stderr.write(f"Empty CSV: {src_path}\n")
        return
    headers = rows[0]
    strip_idx = [i for i, h in enumerate(headers) if h in dob_columns]
    if not strip_idx:
        sys.stderr.write(
            f"WARN: no PII columns to strip in {src_name} (headers: {headers})\n"
        )
    with open(dst_path, "w", newline="", encoding="utf-8") as fout:
        writer = csv.writer(fout, lineterminator="\n")
        writer.writerow(headers)
        for row in rows[1:]:
            for i in strip_idx:
                if i < len(row):
                    row[i] = ""
            writer.writerow(row)
    print(f"wrote {dst_path} (stripped columns: {[headers[i] for i in strip_idx]})")


def anonymise_xlsx(src_name: str, dst_name: str, dob_columns_per_sheet: dict) -> None:
    """Copy xlsx from src → dst, blanking any cell whose header (row 1) matches dob_columns."""
    src_path = SRC / src_name
    dst_path = OUT / dst_name
    wb = openpyxl.load_workbook(src_path, data_only=False)
    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]
        dob_cols = dob_columns_per_sheet.get(sheet_name, [])
        if not dob_cols:
            continue
        # find column indices by header (row 1)
        header_row = [c.value for c in sh[1]]
        strip_idx_zero_based = [i for i, h in enumerate(header_row) if h in dob_cols]
        if not strip_idx_zero_based:
            sys.stderr.write(
                f"WARN: no PII columns to strip in sheet '{sheet_name}' of {src_name} "
                f"(headers row 1: {header_row})\n"
            )
            continue
        # iter rows from row 2; openpyxl is 1-indexed
        for row in sh.iter_rows(min_row=2):
            for i in strip_idx_zero_based:
                if i < len(row):
                    row[i].value = None
        print(
            f"  sheet '{sheet_name}': stripped {len(strip_idx_zero_based)} column(s) "
            f"({[header_row[i] for i in strip_idx_zero_based]})"
        )
    wb.save(dst_path)
    print(f"wrote {dst_path}")


def main() -> None:
    # 1. PayHistory CSV — no PII columns to strip; copy through.
    anonymise_csv("Virtus SAMPLE PayHistorySampleFile(in).csv", "virtus-payhistory.csv", [])
    # 2. PayRateHistory CSV — no PII columns to strip; copy through.
    anonymise_csv(
        "Virtus SAMPLE PayRateHistorySampleFile(in).csv", "virtus-payratehistory.csv", []
    )
    # 3. PositionHistory CSV — strip Employee DOB.
    anonymise_csv(
        "Virtus SAMPLE PositionHistorySampleFile(in).csv",
        "virtus-positionhistory.csv",
        ["Employee DOB"],
    )
    # 4. 3-sheet workbook — strip Employee DOB on Sheet2.
    anonymise_xlsx(
        "Virtus Health LSL - Sample run.xlsx",
        "virtus-3sheet.xlsx",
        {"Sheet2": ["Employee DOB"]},
    )


if __name__ == "__main__":
    main()
